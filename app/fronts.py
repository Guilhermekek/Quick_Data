"""
Detecção e importação de "Fronts" (abas de relatório/layout) de arquivos
Quick Data externos — equivalente ao Form_Importacao do VBA legado
(RN-101 a RN-104 na documentação técnica).

Abas de infraestrutura (Base, tabelas mestre, config, log) são excluídas
da lista por padrão; o restante é oferecido ao usuário para importar.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Nomes de aba que fazem parte da infraestrutura do Quick Data, não são
# "Fronts" de relatório. Comparação sem diferenciar maiúscula/minúscula.
SUPPORT_SHEET_NAMES = {
    "front >>", "control panel >>", "support >>", "aux_extracoes >>",
    "extracao", "ajustes", "base",
    "ref_cruzada_1", "ref_cruzada_2", "sup_linhas", "aux_ifrs16",
    "valid_lin", "cc bd", "dropcomb", "visao_it", "aux",
    "tk_lista_de_erros", "rgm_1", "rgm_2", "fixed_1", "fixed_2",
    "mockup_rgm", "dp_segmento", "listdefinednames", "lista_arq_aux",
    "painel_dm", "dp_rateio",
}


@dataclass
class SheetInfo:
    name: str
    is_front: bool
    rows: int
    cols: int


def _is_front(sheet_name: str) -> bool:
    key = sheet_name.strip().lower()
    if key in SUPPORT_SHEET_NAMES:
        return False
    if key.endswith(">>"):
        return False
    return True


def list_sheets(path: str) -> list[SheetInfo]:
    """Abre o arquivo (somente leitura) e devolve todas as abas, marcando
    quais parecem ser Fronts (candidatas a importação) vs. infraestrutura."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        result = []
        for name in wb.sheetnames:
            ws = wb[name]
            result.append(
                SheetInfo(
                    name=name,
                    is_front=_is_front(name),
                    rows=ws.max_row or 0,
                    cols=ws.max_column or 0,
                )
            )
        return result
    finally:
        wb.close()


_SUFFIX_RE = re.compile(r"\s\((\d+)\)$")


def _unique_sheet_name(wb: openpyxl.Workbook, desired: str) -> str:
    """Replica a regra do Form_Importacao original: nome duplicado ganha
    sufixo numérico incremental em vez de travar a importação (RN-103).

    Normaliza um sufixo " (N)" já existente em `desired` antes de checar
    colisão — sem isso, reimportar a partir de um arquivo que já tem
    "Adicoes (2)" (por exemplo, reusando o próprio destino como origem
    numa sessão de teste) empilhava sufixos: "Adicoes (2) (2)".
    """
    desired = _SUFFIX_RE.sub("", desired).strip() or desired
    if desired not in wb.sheetnames:
        return desired
    n = 2
    while f"{desired} ({n})" in wb.sheetnames:
        n += 1
    return f"{desired} ({n})"


def _json_safe(value):
    """openpyxl devolve datetime/Decimal/etc. para algumas células — troca
    qualquer coisa que não seja um tipo JSON nativo pela representação em
    texto, só para exibição no viewer (não é usado para gravar arquivo)."""
    if value is None or isinstance(value, (int, float, str, bool)):
        return value
    return str(value)


def get_sheet_data(path: str, sheet_name: str, start_row: int = 1, max_rows: int = 40) -> dict:
    """Lê uma página de linhas de uma aba para o visualizador do app —
    a função que faltava para a Importação de Front deixar de ser só
    "salva um arquivo novo" e passar a mostrar o conteúdo dentro do
    próprio Quick Data, como a planilha original sempre fez (lá a aba
    importada é uma aba de Excel de verdade, então "ver" era só olhar
    pra ela). Pagina em vez de carregar a aba inteira de uma vez — Fronts
    reais chegam a ter centenas de colunas e milhares de linhas.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f'Aba "{sheet_name}" não existe em {Path(path).name}')
        ws = wb[sheet_name]
        total_rows = ws.max_row or 0
        total_cols = ws.max_column or 0
        start_row = max(1, start_row)
        end_row = min(start_row + max_rows - 1, total_rows)

        rows: list[list] = []
        if total_rows and total_cols and start_row <= total_rows:
            for row in ws.iter_rows(min_row=start_row, max_row=end_row, max_col=total_cols):
                rows.append([_json_safe(cell.value) for cell in row])

        return {
            "sheetName": sheet_name,
            "totalRows": total_rows,
            "totalCols": total_cols,
            "startRow": start_row,
            "rows": rows,
        }
    finally:
        wb.close()


def apply_sheet_edits(
    path: str,
    sheet_name: str,
    cell_edits: list[dict],
    row_colors: list[dict],
    row_formats: list[dict] | None = None,
    column_widths: list[dict] | None = None,
    cell_formats: list[dict] | None = None,
) -> None:
    """Grava direto na aba real do arquivo — é o que faz "editar o Front
    no app" ser edição de verdade, não só um estado que se perde ao virar
    de página. Abre em modo de escrita (sem `read_only`).

    `cell_edits`: [{"row": int, "col": int, "value": ...}, ...]
    `row_colors`: [{"row": int, "color": "RRGGBB" ou None (limpa a cor)}, ...]
    `row_formats`: [{"row": int, "fontFamily": str|None, "fontSize": int|None,
        "height": float|None}, ...] — fonte/tamanho aplicados em todas as
        células da linha (mesmo alcance de colunas que row_colors); altura
        vai em `row_dimensions` (unidade nativa do Excel, pontos).
    `column_widths`: [{"col": int, "width": float|None}, ...] — largura em
        `column_dimensions` (unidade nativa do Excel, "caracteres").
    `cell_formats`: [{"row": int, "col": int, "color": str|None,
        "fontFamily": str|None, "fontSize": int|None}, ...] — mesma ideia
        de row_colors/row_formats, mas restrito a UMA célula (o usuário
        pode formatar a linha inteira OU só uma célula específica dentro
        dela; célula tem prioridade visual sobre a linha, mas ambas são
        formatação real do Excel, não algo cosmético só na tela).
    """
    wb = openpyxl.load_workbook(path)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f'Aba "{sheet_name}" não existe em {Path(path).name}')
        ws = wb[sheet_name]

        for edit in cell_edits:
            ws.cell(row=edit["row"], column=edit["col"], value=edit["value"])

        max_col = max(ws.max_column or 1, 1)
        for rc in row_colors:
            color = rc.get("color")
            fill = (
                PatternFill(start_color=color, end_color=color, fill_type="solid")
                if color
                else PatternFill(fill_type=None)
            )
            for col in range(1, max_col + 1):
                ws.cell(row=rc["row"], column=col).fill = fill

        for rf in row_formats or []:
            row = rf["row"]
            family = rf.get("fontFamily")
            size = rf.get("fontSize")
            if family or size:
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    base = cell.font
                    cell.font = Font(
                        name=family or base.name,
                        size=size or base.size,
                        bold=base.bold, italic=base.italic, color=base.color,
                    )
            height = rf.get("height")
            if height is not None:
                ws.row_dimensions[row].height = height or None

        for cw in column_widths or []:
            letter = get_column_letter(cw["col"])
            width = cw.get("width")
            ws.column_dimensions[letter].width = width or None

        for cf in cell_formats or []:
            cell = ws.cell(row=cf["row"], column=cf["col"])
            color = cf.get("color")
            cell.fill = (
                PatternFill(start_color=color, end_color=color, fill_type="solid")
                if color
                else PatternFill(fill_type=None)
            )
            family = cf.get("fontFamily")
            size = cf.get("fontSize")
            if family or size:
                base = cell.font
                cell.font = Font(
                    name=family or base.name,
                    size=size or base.size,
                    bold=base.bold, italic=base.italic, color=base.color,
                )

        wb.save(path)
    finally:
        wb.close()


def delete_sheet_column(path: str, sheet_name: str, col: int) -> None:
    """Remove uma coluna inteira da aba — desloca as colunas seguintes
    para a esquerda em TODAS as linhas da planilha (igual ao Excel),
    não só na página visível no momento. `col` é 1-based (A=1, B=2, ...).
    """
    wb = openpyxl.load_workbook(path)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f'Aba "{sheet_name}" não existe em {Path(path).name}')
        ws = wb[sheet_name]
        ws.delete_cols(col, 1)
        wb.save(path)
    finally:
        wb.close()


def insert_sheet_column(path: str, sheet_name: str, col: int) -> None:
    """Insere uma coluna em branco na aba real — desloca as colunas a
    partir de `col` para a direita em TODAS as linhas da planilha.
    `col` é 1-based; para acrescentar no fim, use totalCols + 1.

    Escreve uma célula vazia na primeira linha da coluna nova: sem isso
    o openpyxl não "registra" a coluna (nenhuma célula com valor nela),
    e ela não apareceria no viewer, que usa `max_column` pra saber
    quantas colunas existem.
    """
    wb = openpyxl.load_workbook(path)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f'Aba "{sheet_name}" não existe em {Path(path).name}')
        ws = wb[sheet_name]
        ws.insert_cols(col, 1)
        ws.cell(row=1, column=col, value="")
        wb.save(path)
    finally:
        wb.close()


def build_output(source_path: str, items: list[dict], dest_path: str) -> list[str]:
    """Monta `dest_path` a partir de uma lista ORDENADA de itens — a ordem
    de `items` é a ordem final das abas no destino (o usuário define isso
    arrastando na tela). Cada item é:
      - {"type": "front", "name": ...} — copia a aba de `source_path`.
        Só valores + largura de coluna + formato numérico são preservados
        (fórmulas não são copiadas, mesma simplificação que a exportação
        de Front do VBA original já fazia).
      - {"type": "new", "name": ..., "columns": [...]} — cria uma aba em
        branco com esses títulos de coluna na primeira linha (botão
        "Nova Aba" da tela de importação/exportação).

    Acrescenta a um `dest_path` já existente, como o Form_Importacao
    original fazia, em vez de sempre sobrescrever do zero.

    Retorna a lista de nomes finais das abas criadas no destino.
    """
    src = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    dest_file = Path(dest_path)
    if dest_file.exists():
        dest = openpyxl.load_workbook(dest_path)
    else:
        dest = openpyxl.Workbook()
        # remove a aba padrão em branco que o openpyxl cria sozinho
        default = dest.active
        dest.remove(default)

    created: list[str] = []
    try:
        for item in items:
            desired = (item.get("name") or "Nova aba").strip() or "Nova aba"
            final_name = _unique_sheet_name(dest, desired[:31])  # limite do Excel

            if item.get("type") == "new":
                dst_ws = dest.create_sheet(title=final_name)
                for col_idx, header in enumerate(item.get("columns") or [], start=1):
                    cell = dst_ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
            else:
                if desired not in src.sheetnames:
                    continue
                src_ws = src[desired]
                dst_ws = dest.create_sheet(title=final_name)

                for row in src_ws.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue
                        dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.number_format:
                            dst_ws.cell(row=cell.row, column=cell.column).number_format = cell.number_format

                for col_idx, dim in getattr(src_ws, "column_dimensions", {}).items():
                    if dim.width:
                        dst_ws.column_dimensions[col_idx].width = dim.width

            created.append(final_name)

        dest.save(dest_path)
    finally:
        src.close()
        dest.close()

    return created
